# -*- coding: utf-8 -*-
'''
@File : write_xlsx.py
@Time : 2025/06/04 10:54:13
@Author : wlj 
'''

import time
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill, Border, Side
# from openpyxl.chart import PieChart, Reference, ScatterChart, Series
# from openpyxl.chart.label import DataLabelList


def set_style(stylename, fontName='Microsoft YaHei',bold=False,italic=False, horizontal='center', 
              color='000000', fill=False, fillcolor="B8F0C1", front_size=10, border=False):# little green by fault
    my_style = NamedStyle(name=stylename)
    # 'Times New Roman', '黑体'
    font = Font(name=fontName, italic=italic, bold=bold, color=color, size=front_size)  # color='000000'=black, size=12
    # 设置水平居中, 设置垂直居中, 设置后遇到换行符将自动换行
    alignment = Alignment(horizontal=horizontal, vertical='center', wrapText=True)
    # wrapText=True:设置自动换行
    my_style.font = font
    my_style.alignment = alignment

    if fill:
        my_style.fill = PatternFill(fill_type="solid", fgColor=fillcolor) 
    if border:
        side = Side(style='thin', color='000000') # 'thin','thick', 'dashed', '000000':black
        my_style.border = Border(left=side, right=side, top=side, bottom=side)
    return my_style


def write_xlsx_preface(mysheet, style, start_row=1, max_column=9):
    data = "文本数据量"
    key_count = "关键词数量"
    key_merge = "关键词是否合并"
    line_process = "文本处理方式\n" \
                   "0: 整体处理   1: 逐行处理"
    match_method = "匹配条件"
    match_case = "匹配大小写"
    match_wholeword = "匹配全词"
    time_elapse = "耗时(S)"
    # memory_elapse = "内存消耗"
    extral = "备注"
    test_time = "测试时间: "

    test_time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    mysheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_column) # 
    mysheet.cell(start_row, 1, value=f"{test_time} : {test_time_str}").style = style   

    start_row += 1
    end_row = start_row + 1  
    
    mysheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1) # 
    mysheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2) # 
    mysheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3) # 
    mysheet.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4) # 
    mysheet.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=6) # 
    mysheet.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7) # 
    mysheet.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8) # 
    mysheet.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9) # 
    

    mysheet.cell(start_row, 1, value=data).style = style      
    mysheet.cell(start_row, 2, value=key_count).style = style
    mysheet.cell(start_row, 3, value=key_merge).style = style      
    mysheet.cell(start_row, 4, value=line_process).style = style
    mysheet.cell(start_row, 5, value=match_method).style = style   
    mysheet.cell(start_row + 1, 5, value=match_case).style = style   
    mysheet.cell(start_row + 1, 6, value=match_wholeword).style = style      
    mysheet.cell(start_row, 7, value=time_elapse).style = style     
    # mysheet.cell(start_row, 8, value=memory_elapse).style = style
    mysheet.cell(start_row, 8, value=extral).style = style
    
    return mysheet, end_row

'''
******************************************************************
args:
    configs = {'match_case': match_case,  # 匹配大小写
               'match_wholeword': match_wholeword,  # 全词匹配
               'line_by_line': line_by_line,  # 逐行处理
               'keyword_merge': keyword_merge,  # 关键词是否合并
              }
    keywords:  a list of key words
return:
    
******************************************************************
'''
def write_content(mysheet, style, configs, key_words, time_elaspe, row=1):
    
    mysheet.cell(row, 2, value=len(key_words)).style = style
    mysheet.cell(row, 3, value=configs['keyword_merge']).style = style
    mysheet.cell(row, 4, value=configs['line_by_line']).style = style
    mysheet.cell(row, 5, value=configs['match_case']).style = style
    mysheet.cell(row, 6, value=configs['match_wholeword']).style = style
    mysheet.cell(row, 7, value="%.2f"%time_elaspe).style = style

    return mysheet
  

def write_xlsx_llm_preface(mysheet, style, start_row=1, max_column=8):
    engine = "Inference Engine"
    model = "Model"
    prompt = "prompt(输入提示词)"
    input_len = " Input Length\n(输入长度, token)"
    quantization = "Quantization"
    metric = "metirc"
    memory = "GPU Memory (MB)"
    model_memory = "模型内存"
    total_memory = "总内存"
    # memory_elapse = "内存消耗"
    extral = "备注"
    test_time = "测试时间: "

    test_time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    mysheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_column) # 
    mysheet.cell(start_row, 1, value=f"{test_time} : {test_time_str}").style = style   

    start_row += 1
    end_row = start_row + 1

    mysheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1) # 
    mysheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2) # 
    mysheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3) # 
    mysheet.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4) # 
    mysheet.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5) # 
    mysheet.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6) # 

    mysheet.merge_cells(start_row=start_row, start_column=7, end_row=start_row, end_column=8) # 

    mysheet.cell(start_row, 1, value=engine).style = style      
    mysheet.cell(start_row, 2, value=model).style = style
    mysheet.cell(start_row, 3, value=prompt).style = style      
    mysheet.cell(start_row, 4, value=input_len).style = style
    mysheet.cell(start_row, 5, value=quantization).style = style   
    mysheet.cell(start_row, 6, value=metric).style = style   

    mysheet.cell(start_row, 7, value=memory).style = style   
    mysheet.cell(end_row, 7, value=model_memory).style = style   
    mysheet.cell(end_row, 8, value=total_memory).style = style   

    return mysheet, end_row

def write_llm_content(mysheet, style, configs, row=1):
    
    mysheet.cell(row, 2, value=configs['model']).style = style
    mysheet.cell(row, 3, value=configs['prompt']).style = style
    mysheet.cell(row, 4, value=configs['input_length']).style = style
    mysheet.cell(row, 5, value=configs['quantization']).style = style
    mysheet.cell(row, 6, value=configs['metric']).style = style
    mysheet.cell(row, 7, value=configs['model_memory']).style = style
    mysheet.cell(row, 8, value=configs['total_memory']).style = style

    return mysheet
if __name__ == '__main__':
    a = 0
    

